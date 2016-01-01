# =HYPERLINK("http://www.amazon.com/gp/search/ref=sr_adv_b?search-alias=stripbooks&unfiltered=1&field-isbn="&A2&"&sort=
from openpyxl import load_workbook
import time, sys, webbrowser, os

folder_ISBN_Incoming = "ISBN_Incoming"
folder_ISBN_Processed= "ISBN_Processed"
folder_ExcelFiles = "ExcelFiles"

# process each file in the "Incoming" folder
curdir = os.getcwd()
for root, dirs, files in os.walk(curdir + '\\' + folder_ISBN_Incoming, topdown=False):
    for name in files:
        if ".txt" in name:
            xl_template = load_workbook('book-listing-template.xltm')
            ws = xl_template.active
            filename = os.path.join(root, name)
            lines = [line.rstrip('\n') for line in open(filename)]
            # output_filename = curdir + '\test\' + str(name) + '.xlsx'
            # insert the ISBN numbers
            lineNum=2
            for isbn in lines:
                print('Processing ISBN: ' + isbn)
                ws['A' + str(lineNum)] = isbn
                amazon_booklist_url = "http://www.amazon.com/gp/search/ref=sr_adv_b/?search-alias=stripbooks&unfiltered=1&field-isbn=" + isbn + "&sort=relevanceexprank"
                formula = '=HYPERLINK("' + amazon_booklist_url + '","Amazon Link")'	
                ws['B' + str(lineNum)] = formula
                #print(formula)
                # Open a browser tab for each book
                #webbrowser.open(amazon_booklist_url)
                
                lineNum += 1
        
            # save the new excel file	
            xl_template.save(folder_ExcelFiles + "\\" + name.rstrip('.txt')  + "_" + time.strftime("%Y-%m-%d_%H-%M-%S") + '.xlsx')
            # Move the incoming text file to the processed folder
            os.rename(filename, folder_ISBN_Processed + "\\" + name)
