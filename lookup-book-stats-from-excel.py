# =HYPERLINK("http://www.amazon.com/gp/search/ref=sr_adv_b?search-alias=stripbooks&unfiltered=1&field-isbn="&A2&"&sort=
from openpyxl import load_workbook
import time, sys, webbrowser

	
lines = [line.rstrip('\n') for line in open('sample-isbn-numbers.txt')]
	
	
# create the new excel workbook
xl_template = load_workbook('book-listing-template.xltm')
ws = xl_template.active

# insert the ISBN numbers
lineNum=2
for isbn in lines:
	print('Processing ISBN: ' + isbn)
	ws['A' + str(lineNum)] = isbn
	amazon_booklist_url = "http://www.amazon.com/gp/search/ref=sr_adv_b/?search-alias=stripbooks&unfiltered=1&field-isbn=" + isbn + "&sort=relevanceexprank"
	formula = '=HYPERLINK("' + amazon_booklist_url + '","Amazon Link")'	
	ws['B' + str(lineNum)] = formula
	print(formula)
	# Open a browser tab for each book
	webbrowser.open(amazon_booklist_url)
	
	lineNum += 1



	
# save the new excel file	
xl_template.save('test/book-list-' + time.strftime("%Y-%m-%d_%H-%M-%S") + '.xlsx')